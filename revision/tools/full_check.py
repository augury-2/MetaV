import openpyxl, statistics as st, itertools, math, sys
from collections import Counter

PATH = sys.argv[1] if len(sys.argv) > 1 else '/projects/sandbox/MetaV/Original Data MSME.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]
hdr = [c.value for c in ws[1]]
rows = [dict(zip(hdr, [ws.cell(row=r, column=c).value for c in range(1, len(hdr) + 1)]))
        for r in range(2, ws.max_row + 1)]
rows = [r for r in rows if r.get('respondent_id') is not None]
n = len(rows)

CONS = {'GSCI': 5, 'RP': 4, 'DI': 5, 'TMC': 4, 'SP': 4, 'CBI': 4, 'KSR': 4, 'GLP': 4}
IT = {k: [f"{k}{i}" for i in range(1, m + 1)] for k, m in CONS.items()}
ITEMS = [i for k in IT for i in IT[k]]
V = {i: [r[i] for r in rows] for i in ITEMS}


def corr(a, b):
    ma, mb = st.mean(a), st.mean(b)
    num = sum((x - ma) * (y - mb) for x, y in zip(a, b))
    da = math.sqrt(sum((x - ma) ** 2 for x in a)); db = math.sqrt(sum((y - mb) ** 2 for y in b))
    return num / (da * db) if da and db else 0.0


def sc(r, k):
    return sum(r[f"{k}{i}"] for i in range(1, CONS[k] + 1)) / CONS[k]


S = {k: [sc(r, k) for r in rows] for k in CONS}
BAR = "=" * 76

print(BAR); print(f"FILE: {PATH.split('/')[-1]}"); print(f"ROWS WITH DATA: {n}   (manuscript: 220)"); print(BAR)

g = Counter(r['group'] for r in rows)
print("\n[1] GROUP SIZES        expected developing 100, emerging 120")
for k, v in g.items():
    print(f"      {str(k):12s} {v:4d}")

print("\n[2] INTEGRITY")
miss = sum(1 for r in rows for i in ITEMS if r[i] is None or r[i] == -99)
bad = [(i, r['respondent_id'], r[i]) for r in rows for i in ITEMS
       if r[i] is not None and (not isinstance(r[i], (int, float)) or not 1 <= r[i] <= 5)]
print(f"      missing/-99 {miss}    out-of-range {len(bad)}")
print(f"      value counts {dict(sorted(Counter(r[i] for r in rows for i in ITEMS).items()))}")

print("\n[3] TABLE 3 MARGINALS")
exp3 = {'role': {'logistics_manager': 84, 'sc_director': 61, 'ops_sustainability': 52, 'other_senior': 23},
        'firm_size': {'small': 47, 'medium': 96, 'large': 77},
        'experience': {'lt5': 38, '5to10': 92, 'gt10': 90}}
ok3 = True
for var, exp in exp3.items():
    got = Counter(r[var] for r in rows)
    line = "  ".join(f"{k}={got.get(k,0)}/{exp[k]}" for k in exp)
    m = all(got.get(k, 0) == exp[k] for k in exp)
    ok3 &= m
    print(f"      {var:11s} {line}   {'MATCH' if m else 'MISMATCH'}")

print("\n[4] TABLE 8 MEANS (obs vs pub)")
exp8 = {'GSCI': (3.05, 3.98, 3.56), 'RP': (3.62, 4.18, 3.93), 'DI': (2.58, 4.07, 3.39),
        'TMC': (3.01, 4.31, 3.72), 'SP': (2.88, 3.92, 3.45), 'CBI': (3.00, 3.79, 3.43),
        'KSR': (2.79, 4.02, 3.46), 'GLP': (3.04, 4.21, 3.68)}
dev = [r for r in rows if r['group'] == 'developing']; eme = [r for r in rows if r['group'] == 'emerging']
worst = 0
for k in CONS:
    o = [st.mean([sc(r, k) for r in grp]) for grp in (dev, eme, rows)]
    e = exp8[k]
    d = max(abs(a - b) for a, b in zip(o, e)); worst = max(worst, d)
    print(f"      {k:5s} dev {o[0]:.2f}/{e[0]:.2f}   eme {o[1]:.2f}/{e[1]:.2f}   pool {o[2]:.2f}/{e[2]:.2f}   maxdiff {d:+.2f}")
print(f"      largest deviation across all cells: {worst:.2f}")

mono = {k: st.mean([corr(V[a], V[b]) for a, b in itertools.combinations(IT[k], 2)]) for k in CONS}

print("\n[5] TABLE 4 MEASUREMENT (PC1 approx)   pub: loadings .70-.88, CR .86-.91, AVE .61-.71")


def pc1(items):
    k = len(items)
    mu = [st.mean(V[i]) for i in items]; sd = [st.pstdev(V[i]) for i in items]
    Z = [[(V[items[j]][t] - mu[j]) / sd[j] for j in range(k)] for t in range(n)]
    R = [[sum(Z[t][i] * Z[t][j] for t in range(n)) / n for j in range(k)] for i in range(k)]
    v = [1.0] * k
    for _ in range(800):
        w = [sum(R[i][j] * v[j] for j in range(k)) for i in range(k)]
        nn = math.sqrt(sum(x * x for x in w)); v = [x / nn for x in w]
    lam = math.sqrt(sum(R[i][j] * v[i] * v[j] for i in range(k) for j in range(k)))
    return [x * lam for x in v]


exp4 = {'GSCI': (.89, .62), 'RP': (.88, .65), 'DI': (.90, .64), 'TMC': (.90, .69),
        'SP': (.86, .61), 'CBI': (.87, .63), 'KSR': (.89, .67), 'GLP': (.91, .71)}
for k in CONS:
    L = pc1(IT[k]); ave = sum(l * l for l in L) / len(L)
    s = sum(L); e = sum(1 - l * l for l in L); cr = s * s / (s * s + e)
    a = CONS[k] * mono[k] / (1 + (CONS[k] - 1) * mono[k])
    print(f"      {k:5s} load {min(L):.2f}-{max(L):.2f}   CR {cr:.3f}/{exp4[k][0]:.2f}   AVE {ave:.3f}/{exp4[k][1]:.2f}   alpha {a:.3f}")

print("\n[6] TABLE 5 HTMT (exact)   pub: all below 0.85")
PUB5 = {('RP','GSCI'):.41,('DI','GSCI'):.58,('TMC','GSCI'):.61,('SP','GSCI'):.44,('CBI','GSCI'):.38,
        ('KSR','GSCI'):.49,('DI','RP'):.39,('TMC','RP'):.52,('SP','RP'):.47,('CBI','RP'):.33,
        ('KSR','RP'):.35,('TMC','DI'):.55,('SP','DI'):.42,('CBI','DI'):.46,('KSR','DI'):.63,
        ('SP','TMC'):.49,('CBI','TMC'):.40,('KSR','TMC'):.51,('CBI','SP'):.36,('KSR','SP'):.41,
        ('KSR','CBI'):.44}
mx = 0; br = []
for (a, b), pub in PUB5.items():
    het = st.mean([corr(V[x], V[y]) for x in IT[a] for y in IT[b]])
    h = het / math.sqrt(mono[a] * mono[b]); mx = max(mx, h)
    if h >= .85: br.append((a, b, h))
    print(f"      {a+'-'+b:14s} {h:.3f} / {pub:.2f}   diff {h-pub:+.3f}{'   >>> BREACH' if h>=.85 else ''}")
print(f"      max HTMT {mx:.3f}   breaches of 0.85: {len(br)}")

print("\n[7] HARMAN SINGLE FACTOR   manuscript: 'well under half'; threshold 50%")
K = len(ITEMS)
mu = [st.mean(V[i]) for i in ITEMS]; sd = [st.pstdev(V[i]) for i in ITEMS]
Z = [[(V[ITEMS[j]][t] - mu[j]) / sd[j] for j in range(K)] for t in range(n)]
R = [[sum(Z[t][i] * Z[t][j] for t in range(n)) / n for j in range(K)] for i in range(K)]
Rw = [r[:] for r in R]; ev = []
for _ in range(3):
    v = [1.0] * K
    for _ in range(2000):
        w = [sum(Rw[i][j] * v[j] for j in range(K)) for i in range(K)]
        nn = math.sqrt(sum(x * x for x in w))
        if nn == 0: break
        v = [x / nn for x in w]
    lam = sum(Rw[i][j] * v[i] * v[j] for i in range(K) for j in range(K)); ev.append(lam)
    for i in range(K):
        for j in range(K): Rw[i][j] -= lam * v[i] * v[j]
for i, l in enumerate(ev, 1):
    print(f"      factor {i}: {l/K*100:6.2f}%")
print(f"      >>> FIRST FACTOR {ev[0]/K*100:.2f}%")


def ols(y, Xs):
    m = len(Xs); N = len(y); cols = [[1.0] * N] + [list(x) for x in Xs]; M = m + 1
    A = [[sum(cols[i][t] * cols[j][t] for t in range(N)) for j in range(M)] for i in range(M)]
    b = [sum(cols[i][t] * y[t] for t in range(N)) for i in range(M)]
    for i in range(M):
        p = max(range(i, M), key=lambda r: abs(A[r][i])); A[i], A[p] = A[p], A[i]; b[i], b[p] = b[p], b[i]
        d = A[i][i]; A[i] = [x / d for x in A[i]]; b[i] /= d
        for r2 in range(M):
            if r2 != i and A[r2][i]:
                f = A[r2][i]; A[r2] = [x - f * c for x, c in zip(A[r2], A[i])]; b[r2] -= f * b[i]
    yh = [b[0] + sum(b[j + 1] * Xs[j][t] for j in range(m)) for t in range(N)]
    ym = st.mean(y); ss = sum((v - ym) ** 2 for v in y)
    return b, 1 - sum((v - h) ** 2 for v, h in zip(y, yh)) / ss


def z(v):
    m, s = st.mean(v), st.pstdev(v); return [(x - m) / s for x in v]


print("\n[8] FULL COLLINEARITY VIF   threshold 3.3")
for k in CONS:
    _, r2 = ols(S[k], [S[o] for o in CONS if o != k])
    vif = 1 / (1 - r2)
    print(f"      {k:5s} VIF {vif:6.3f}{'   >>> BREACH' if vif > 3.3 else ''}")

print("\n[9] TABLE 6 PATHS (OLS on composites, indicative)")
spec = [('TMC', ['RP', 'SP'], {'RP': .44, 'SP': .29}),
        ('GSCI', ['TMC', 'DI', 'CBI'], {'TMC': .38, 'DI': .41, 'CBI': .16}),
        ('DI', ['KSR'], {'KSR': .61}),
        ('GLP', ['GSCI', 'TMC'], {'GSCI': .49, 'TMC': .27})]
for dv, ivs, pub in spec:
    b, r2 = ols(z(S[dv]), [z(S[i]) for i in ivs])
    print(f"      -> {dv:5s}  R2 {r2:.3f}")
    for j, iv in enumerate(ivs):
        print(f"           {iv:5s} {b[j+1]:+.3f} / {pub[iv]:+.2f}")

print("\n[10] OTHER COLUMNS")
for v in ('country', 'sector', 'ems'):
    print(f"      {v}: {dict(Counter(str(r[v]) for r in rows))}")
ct = Counter((r['country'], r['group']) for r in rows)
print(f"      country x group: {dict(ct)}")
