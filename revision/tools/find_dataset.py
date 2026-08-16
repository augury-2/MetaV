#!/usr/bin/env python3
"""
find_dataset.py - identify which file on a data drive matches the published manuscript.

Walks a folder, opens every .xlsx / .xls / .csv it finds, tries to locate the 34
green-logistics items, and scores each file against the benchmarks published in
"Enablers of Green Logistics Adoption in Developing versus Emerging Economies".

Files are ranked by how many benchmarks they reproduce, so the correct dataset
should appear at the top even when many files share the same name.

USAGE
    pip install openpyxl
    python3 find_dataset.py "/path/to/data collection drive"

Nothing is modified. The script only reads.
"""
import sys, os, csv, math, statistics as st, itertools
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl first:  pip install openpyxl")

CONS = {'GSCI': 5, 'RP': 4, 'DI': 5, 'TMC': 4, 'SP': 4, 'CBI': 4, 'KSR': 4, 'GLP': 4}
IT = {k: [f"{k}{i}" for i in range(1, m + 1)] for k, m in CONS.items()}
ITEMS = [i for k in IT for i in IT[k]]

# ---- published benchmarks ---------------------------------------------------
PUB_N = 220
PUB_GROUPS = {'developing': 100, 'emerging': 120}
PUB_MARGINALS = {'role': {'logistics_manager': 84, 'sc_director': 61,
                          'ops_sustainability': 52, 'other_senior': 23},
                 'firm_size': {'small': 47, 'medium': 96, 'large': 77},
                 'experience': {'lt5': 38, '5to10': 92, 'gt10': 90}}
PUB_T8 = {'GSCI': (3.05, 3.98, 3.56), 'RP': (3.62, 4.18, 3.93), 'DI': (2.58, 4.07, 3.39),
          'TMC': (3.01, 4.31, 3.72), 'SP': (2.88, 3.92, 3.45), 'CBI': (3.00, 3.79, 3.43),
          'KSR': (2.79, 4.02, 3.46), 'GLP': (3.04, 4.21, 3.68)}
PUB_T4 = {'GSCI': (.89, .62), 'RP': (.88, .65), 'DI': (.90, .64), 'TMC': (.90, .69),
          'SP': (.86, .61), 'CBI': (.87, .63), 'KSR': (.89, .67), 'GLP': (.91, .71)}
PUB_T5_RANGE = (0.33, 0.63)          # all published HTMT ratios fall in this band
PUB_HARMAN_MAX = 50.0                # manuscript claims "well under half"
PUB_T6 = {('RP', 'TMC'): .44, ('SP', 'TMC'): .29, ('TMC', 'GSCI'): .38, ('DI', 'GSCI'): .41,
          ('CBI', 'GSCI'): .16, ('KSR', 'DI'): .61, ('GSCI', 'GLP'): .49, ('TMC', 'GLP'): .27}


# ---- helpers ----------------------------------------------------------------
def corr(a, b):
    ma, mb = st.mean(a), st.mean(b)
    num = sum((x - ma) * (y - mb) for x, y in zip(a, b))
    da = math.sqrt(sum((x - ma) ** 2 for x in a))
    db = math.sqrt(sum((y - mb) ** 2 for y in b))
    return num / (da * db) if da and db else 0.0


def norm(s):
    return ''.join(ch for ch in str(s).upper() if ch.isalnum())


def read_tables(path):
    """Yield (label, header_list, list_of_rows) for each sheet or csv."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try:
            with open(path, newline='', encoding='utf-8-sig', errors='replace') as f:
                rr = list(csv.reader(f))
            if rr:
                yield '', rr[0], rr[1:]
        except Exception:
            return
    elif ext in ('.xlsx', '.xlsm'):
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception:
            return
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(rows) > 1:
                yield name, rows[0], rows[1:]
        wb.close()


def extract(header, body):
    """Locate the 34 items. Return (records, colmap) or (None, None)."""
    hmap = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        hmap.setdefault(norm(h), idx)
    cols = {}
    for it in ITEMS:
        if norm(it) in hmap:
            cols[it] = hmap[norm(it)]
    if len(cols) < 34:
        return None, None
    meta = {}
    for m in ('group', 'role', 'firm_size', 'experience', 'country'):
        if norm(m) in hmap:
            meta[m] = hmap[norm(m)]
    recs = []
    for r in body:
        rec = {}
        ok = True
        for it, ci in cols.items():
            v = r[ci] if ci < len(r) else None
            try:
                v = int(float(v))
            except (TypeError, ValueError):
                ok = False
                break
            if not 1 <= v <= 5:
                ok = False
                break
            rec[it] = v
        if not ok:
            continue
        for m, ci in meta.items():
            rec[m] = str(r[ci]).strip().lower() if ci < len(r) and r[ci] is not None else None
        recs.append(rec)
    return (recs, meta) if len(recs) >= 30 else (None, None)


def ols(y, Xs):
    m = len(Xs); N = len(y); cols = [[1.0] * N] + [list(x) for x in Xs]; M = m + 1
    A = [[sum(cols[i][t] * cols[j][t] for t in range(N)) for j in range(M)] for i in range(M)]
    b = [sum(cols[i][t] * y[t] for t in range(N)) for i in range(M)]
    for i in range(M):
        p = max(range(i, M), key=lambda rr: abs(A[rr][i]))
        A[i], A[p] = A[p], A[i]; b[i], b[p] = b[p], b[i]
        if A[i][i] == 0:
            return None
        d = A[i][i]; A[i] = [x / d for x in A[i]]; b[i] /= d
        for r2 in range(M):
            if r2 != i and A[r2][i]:
                f = A[r2][i]
                A[r2] = [x - f * c for x, c in zip(A[r2], A[i])]
                b[r2] -= f * b[i]
    return b


def zs(v):
    m, s = st.mean(v), st.pstdev(v)
    return [(x - m) / s for x in v] if s else [0.0] * len(v)


def pc1(V, items, n):
    k = len(items)
    mu = [st.mean(V[i]) for i in items]; sd = [st.pstdev(V[i]) for i in items]
    if min(sd) == 0:
        return [0.0] * k
    Z = [[(V[items[j]][t] - mu[j]) / sd[j] for j in range(k)] for t in range(n)]
    R = [[sum(Z[t][i] * Z[t][j] for t in range(n)) / n for j in range(k)] for i in range(k)]
    v = [1.0] * k
    for _ in range(400):
        w = [sum(R[i][j] * v[j] for j in range(k)) for i in range(k)]
        nn = math.sqrt(sum(x * x for x in w))
        if nn == 0:
            break
        v = [x / nn for x in w]
    lam = math.sqrt(max(0.0, sum(R[i][j] * v[i] * v[j] for i in range(k) for j in range(k))))
    return [x * lam for x in v]


def assess(recs):
    n = len(recs)
    V = {i: [r[i] for r in recs] for i in ITEMS}
    S = {k: [sum(r[f"{k}{j}"] for j in range(1, CONS[k] + 1)) / CONS[k] for r in recs] for k in CONS}
    res = {'n': n, 'checks': {}}

    res['checks']['n=220'] = (n == PUB_N)

    gc = Counter(r.get('group') for r in recs)
    res['groups'] = dict(gc)
    res['checks']['groups 100/120'] = (gc.get('developing') == 100 and gc.get('emerging') == 120)

    marg_ok = True
    for var, exp in PUB_MARGINALS.items():
        got = Counter(r.get(var) for r in recs)
        if not all(got.get(k, 0) == v for k, v in exp.items()):
            marg_ok = False
    res['checks']['Table 3 marginals'] = marg_ok

    dev = [r for r in recs if r.get('group') == 'developing']
    eme = [r for r in recs if r.get('group') == 'emerging']
    dmax = 0.0
    if dev and eme:
        for k in CONS:
            o = (st.mean([sum(r[f"{k}{j}"] for j in range(1, CONS[k] + 1)) / CONS[k] for r in dev]),
                 st.mean([sum(r[f"{k}{j}"] for j in range(1, CONS[k] + 1)) / CONS[k] for r in eme]),
                 st.mean(S[k]))
            dmax = max(dmax, max(abs(a - b) for a, b in zip(o, PUB_T8[k])))
    else:
        dmax = float('nan')
    res['t8_maxdev'] = dmax
    res['checks']['Table 8 means (<=0.20)'] = (dmax == dmax and dmax <= 0.20)

    mono = {k: st.mean([corr(V[a], V[b]) for a, b in itertools.combinations(IT[k], 2)]) for k in CONS}
    cr_dev = ave_dev = 0.0
    for k in CONS:
        L = pc1(V, IT[k], n)
        ave = sum(l * l for l in L) / len(L)
        s = sum(L); e = sum(1 - l * l for l in L)
        cr = s * s / (s * s + e) if (s * s + e) else 0
        cr_dev = max(cr_dev, abs(cr - PUB_T4[k][0]))
        ave_dev = max(ave_dev, abs(ave - PUB_T4[k][1]))
    res['t4_cr_dev'] = cr_dev; res['t4_ave_dev'] = ave_dev
    res['checks']['Table 4 CR/AVE (<=0.08)'] = (cr_dev <= 0.08 and ave_dev <= 0.08)

    hts = []
    for a, b in itertools.combinations(CONS, 2):
        if mono[a] <= 0 or mono[b] <= 0:
            continue
        het = st.mean([corr(V[x], V[y]) for x in IT[a] for y in IT[b]])
        hts.append(het / math.sqrt(mono[a] * mono[b]))
    res['htmt_min'] = min(hts) if hts else float('nan')
    res['htmt_max'] = max(hts) if hts else float('nan')
    res['checks']['Table 5 HTMT band'] = bool(hts) and \
        res['htmt_min'] >= PUB_T5_RANGE[0] - 0.12 and res['htmt_max'] <= 0.85

    K = len(ITEMS)
    mu = [st.mean(V[i]) for i in ITEMS]; sd = [st.pstdev(V[i]) for i in ITEMS]
    if min(sd) > 0:
        Z = [[(V[ITEMS[j]][t] - mu[j]) / sd[j] for j in range(K)] for t in range(n)]
        R = [[sum(Z[t][i] * Z[t][j] for t in range(n)) / n for j in range(K)] for i in range(K)]
        v = [1.0] * K
        for _ in range(800):
            w = [sum(R[i][j] * v[j] for j in range(K)) for i in range(K)]
            nn = math.sqrt(sum(x * x for x in w))
            if nn == 0:
                break
            v = [x / nn for x in w]
        lam = sum(R[i][j] * v[i] * v[j] for i in range(K) for j in range(K))
        res['harman'] = lam / K * 100
    else:
        res['harman'] = float('nan')
    res['checks']['Harman < 50%'] = (res['harman'] == res['harman'] and res['harman'] < PUB_HARMAN_MAX)

    pmax = 0.0
    for (iv, dv), pub in PUB_T6.items():
        preds = {'TMC': ['RP', 'SP'], 'GSCI': ['TMC', 'DI', 'CBI'], 'DI': ['KSR'],
                 'GLP': ['GSCI', 'TMC']}[dv]
        b = ols(zs(S[dv]), [zs(S[p]) for p in preds])
        if b is None:
            pmax = float('nan'); break
        pmax = max(pmax, abs(b[preds.index(iv) + 1] - pub))
    res['t6_maxdev'] = pmax
    res['checks']['Table 6 paths (<=0.15)'] = (pmax == pmax and pmax <= 0.15)

    res['score'] = sum(1 for v in res['checks'].values() if v)
    return res


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    root = sys.argv[1]
    found = []
    for dp, _, fns in os.walk(root):
        for fn in fns:
            if fn.startswith('~$') or os.path.splitext(fn)[1].lower() not in ('.xlsx', '.xlsm', '.csv'):
                continue
            full = os.path.join(dp, fn)
            for sheet, hdr, body in read_tables(full):
                recs, _ = extract(hdr, body)
                if recs:
                    try:
                        r = assess(recs)
                    except Exception as e:
                        print(f"  ! skipped {fn} [{sheet}]: {e}")
                        continue
                    r['file'] = full; r['sheet'] = sheet
                    found.append(r)

    if not found:
        print("No file containing all 34 items (GSCI1..GLP4) was found under:", root)
        return

    found.sort(key=lambda r: -r['score'])
    names = list(found[0]['checks'].keys())
    print("\n" + "=" * 100)
    print("RANKED CANDIDATES  (8 = reproduces every published benchmark)")
    print("=" * 100)
    for r in found:
        rel = os.path.relpath(r['file'], root)
        print(f"\n[{r['score']}/8]  {rel}" + (f"   sheet '{r['sheet']}'" if r['sheet'] else ""))
        print(f"         n={r['n']}  groups={r['groups']}  Harman={r['harman']:.1f}%  "
              f"HTMT {r['htmt_min']:.2f}-{r['htmt_max']:.2f}")
        print(f"         maxdev  T8={r['t8_maxdev']:.2f}  T4(CR)={r['t4_cr_dev']:.2f}  "
              f"T4(AVE)={r['t4_ave_dev']:.2f}  T6={r['t6_maxdev']:.2f}")
        for k in names:
            print(f"            {'PASS' if r['checks'][k] else 'fail'}  {k}")
    print("\nDone. Send the top-ranked file, or the two highest if none scores 8/8.")


if __name__ == '__main__':
    main()
